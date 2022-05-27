from app import db
from models import User, Pokaz, Counter, House, Address
from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border
from openpyxl.writer.excel import save_virtual_workbook
from datetime import datetime
from utils import month_name


class DataUploader():

    def __init__(self,
                 type: int,
                 mkdid: int = None,
                 kvid: int = None,
                 personid: int = None,
                 period_start: dict = None,
                 period_end: dict = None,
                 user_id: int = None):
        self.__filename = 'Выгрузка_{}_{}.xlsx'.format(datetime.now().strftime("%d-%m-%Y-%H-%M-%S"), user_id)
        self.__user = user_id
        self.mkd_id = mkdid
        self.kv_id = kvid
        self.person_id = personid
        self.type = type
        self.__wb = Workbook()
        self.__ws = self.__wb.active
        self.__ws.title = 'Выгрузка'
        self.period_start_m = period_start['m']
        self.period_end_m = period_end['m']
        self.period_start_y = period_start['y']
        self.period_end_y = period_end['y']
        self.__row = 0
        self.__col = 0
        self.__start_info_row = None
        self.__end_info_row = None

        self.font_bold = Font(name='Arial', size=12, bold=True)
        self.font_regular = Font(name='Arial', size=12)

        self.__cell_filter = 'B6'
        self.__prepareBookHeader()

    def __prepareBookHeader(self):
        cell_sysname = 'A1'
        cell_title = 'A3'
        cell_title_end = 'C3'
        cell_type = 'A4'
        cell_type_in = 'B4'
        cell_date = 'A5'
        cell_date_in = 'B5'
        cell_date_end = 'D5'
        cell_filter_param = 'A6'
        cell_filter_end = 'H6'
        cell_period_param = 'A7'
        cell_period = 'B7'
        cell_period_end = 'D7'

        self.__ws[cell_sysname].font = self.font_bold
        self.__ws[cell_title].font = self.font_bold
        self.__ws[cell_type].font = self.font_bold
        self.__ws[cell_type_in].font = self.font_regular
        self.__ws[cell_date].font = self.font_bold
        self.__ws[cell_date_in].font = self.font_regular
        self.__ws[cell_filter_param].font = self.font_bold
        self.__ws[self.__cell_filter].font = self.font_regular
        self.__ws[cell_period_param].font = self.font_bold
        self.__ws[cell_period].font = self.font_regular

        self.__ws[cell_sysname] = 'АСУПСКУ'
        self.__ws[cell_title] = 'Выгрузка показаний приборов учета'
        self.__ws[cell_date] = 'Дата:'
        self.__ws[cell_date_in] = datetime.now()
        self.__ws[cell_type] = 'Отбор:'
        self.__ws[cell_type_in] = self.__getOtborName()
        self.__ws[cell_date] = 'Дата'
        self.__ws[cell_date_in] = datetime.now().strftime('%d.%m.%Y (%H:%M:%S)')
        self.__ws[cell_filter_param] = 'Фильтр:'
        self.__ws[cell_period_param] = 'Период'
        self.__ws[cell_period] = '{} {} - {} {}'.format(month_name(int(self.period_start_m)),
                                                        self.period_start_y,
                                                        month_name(int(self.period_end_m)),
                                                        self.period_end_y)
        self.__ws.merge_cells('{}:{}'.format(cell_period, cell_period_end))
        self.__ws.merge_cells('{}:{}'.format(cell_date_in, cell_date_end))
        self.__ws.merge_cells('{}:{}'.format(self.__cell_filter, cell_filter_end))
        self.__ws.merge_cells('{}:{}'.format(cell_title,cell_title_end))

        self.__row = 10
        self.__col = 1

    def __getOtborName(self):
        if self.type == 1:
            return 'По МКД'
        elif self.type == 2:
            return 'По квартире'
        elif self.type == 3:
            return 'По пользователю'

    def __getFilterName(self):
        if self.type == 1:
            return 'МКД'
        elif self.type == 2:
            return 'Квартира'
        elif self.type == 3:
            return 'Пользователь'

    def __get_unload_mkd(self):
        mkd = self.mkd_id
        adres = db.session.query(House.adres).filter_by(id=mkd).limit(1).first().adres
        self.__ws[self.__cell_filter] = self.__getFilterName() + ' ' + adres
        self.__prepareUnloadDataTable()
        flats = Address.query.filter_by(house=mkd).all()
        for f in flats:
            if self.__check_exists_pkz(f.id) is False:
                continue
            self.__ws.cell(row=self.__row, column=1, value=f.get_full_address()).font = self.font_bold
            self.__ws.merge_cells(start_row=self.__row,
                                  start_column=1,
                                  end_row=self.__row,
                                  end_column=5)
            self.__ws.cell(row=self.__row, column=1).style = 'Good'

            self.__row += 1
            self.__getUnloadFlat(f.id)

    def __check_exists_pkz(self,flat):
        counters = db.session.query(Counter.id).filter_by(flat=flat).all()
        exists = False
        for c in counters:
            pokaz_data = Pokaz.query.filter(Pokaz.counter == c.id,
                                            Pokaz.p_month >= self.period_start_m,
                                            Pokaz.p_year >= self.period_start_y,
                                            Pokaz.p_month <= self.period_end_m,
                                            Pokaz.p_year <= self.period_end_y).limit(1).first()
            if not pokaz_data:
                continue
            else:
                exists = True
                break
        return exists


    def __getUnloadFlat(self, flat_id=None):
        flat = self.kv_id if flat_id is None else flat_id
        flat_info = Address.query.filter_by(id=flat).limit(1).first()
        flat_adr = flat_info.get_full_address()
        if flat_id is None:
            self.__ws[self.__cell_filter] = self.__getFilterName() + ' ' + flat_adr
            self.__prepareUnloadDataTable()
        flat_counters = db.session.query(Counter.name, Counter.id).filter_by(flat = flat_info.id).all()
        for c in flat_counters:
            pokaz_data = Pokaz.query.filter(Pokaz.counter == c.id,
                                            Pokaz.p_month >= self.period_start_m,
                                            Pokaz.p_year >= self.period_start_y,
                                            Pokaz.p_month <= self.period_end_m,
                                            Pokaz.p_year <= self.period_end_y).order_by(Pokaz.id.desc()).all()
            for p in pokaz_data:
                column = 1
                self.__ws.cell(row=self.__row, column=column, value=c.name)
                column += 1
                self.__ws.cell(row=self.__row, column=column, value='{} {}'.format(month_name(p.p_month),p.p_year))
                column += 1
                self.__ws.cell(row=self.__row, column=column, value=p.amount)
                column += 1
                self.__ws.cell(row=self.__row, column=column, value=p.added_on)
                column += 1
                added_user = db.session.query(User.name).filter_by(id=p.added_by).limit(1).first()
                user_name = added_user.name if added_user else 'Неизвестно'
                self.__ws.cell(row=self.__row, column=column, value=user_name)
                self.__row += 1

    def __prepareUnloadDataTable(self):
        # верхняя строка таблицы
        column = self.__col
        self.__start_info_row = self.__row
        self.__ws.cell(row=self.__row, column=column, value='Прибор учета').font = self.font_bold
        column += 1
        self.__ws.cell(row=self.__row, column=column, value='Период').font = self.font_bold
        column += 1
        self.__ws.cell(row=self.__row, column=column, value='Показания').font = self.font_bold
        column += 1
        self.__ws.cell(row=self.__row, column=column, value='Дата передачи').font = self.font_bold
        column += 1
        self.__ws.cell(row=self.__row, column=column, value='Кем переданы').font = self.font_bold
        for i in range(1, 6):
            self.__ws.cell(row=self.__row, column=i).style = 'Output'
        self.__row += 1

    def __add_bottom(self):
        self.__row += 3
        user = db.session.query(User.name).filter_by(id=self.__user).limit(1).first()
        self.__ws.cell(row=self.__row, column=1, value='Выгрузку запросил: инспектор {}'.format(user.name))

    def __prettify(self):
        for column_cells in self.__ws.columns:
            unmerged_cells = list(
                filter(lambda cell_to_check: cell_to_check.coordinate not in self.__ws.merged_cells, column_cells))
            length = max(len(str(cell.value)) for cell in unmerged_cells)
            self.__ws.column_dimensions[unmerged_cells[0].column_letter].width = length * 1.4
        thin = Side(border_style="thin", color="000000")
        for row in self.__ws.iter_rows(min_row=self.__start_info_row, max_col=5, max_row=self.__end_info_row-1):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def unload_and_push(self):
        if self.type == 2:
            self.__getUnloadFlat()
        elif self.type == 1:
            self.__get_unload_mkd()
        self.__end_info_row = self.__row
        self.__prettify()
        self.__add_bottom()
        return save_virtual_workbook(self.__wb)

    def get_filename(self):
        return self.__filename

    def __download_virtual(self):
        return save_virtual_workbook(self.__wb)